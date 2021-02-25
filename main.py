from flask import Flask , render_template, url_for , session , redirect , request
import mysql.connector
from openpyxl import Workbook,load_workbook
import os
from datetime import date , datetime



#DATABASE KURULUMU
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="sigorta",
)
mycursor = mydb.cursor(dictionary=True)


app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\231xec]/'

@app.route('/' , methods=['GET', 'POST'])
def index():
    if 'id' not in session:
        return redirect(url_for('login'))

    db_name = 'melik'
    if request.method == 'POST':
        db_name = request.form['data']


    mycursor.execute("Show tables;") 
    databasenames = mycursor.fetchall()

    sql = "SELECT * FROM %s ORDER BY id DESC" % (db_name)
    mycursor.execute(sql)
    veriler = mycursor.fetchall()
    return render_template('index.html', veriler = veriler, databasenames=databasenames)
        



@app.route('/login' , methods=['GET' , 'POST'])
def login():
    if 'id' in session:
        return redirect(url_for('index'))
    error = ''
    if request.method == 'POST':
        if request.form['email'] == '':
            error = 'E-posta giriniz!'
        elif request.form['password'] == '':
            error = 'Şifre giriniz!'
        else:
            sql = "SELECT * FROM users WHERE email = %s && password = %s"
            mycursor.execute(sql, (request.form['email'], request.form['password'],))
            user = mycursor.fetchone()
            if user:
                session['id'] = user['id']
                return redirect(url_for('index'))
            else:
                error = 'E-posta veya şifre hatalı'

    return render_template('login.html' , error = error)

@app.route('/police/<url>')
def police(url):
    if 'id' not in session:
        return redirect(url_for('login'))
    sql = "SELECT * FROM aralikkk ORDER BY id DESC"
    mycursor.execute(sql)
    veriler = mycursor.fetchall()
    return render_template('police.html', url=url , veriler=veriler)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/upload', methods = ['GET', 'POST'])
def upload():
    
    if 'id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':

        file = request.files['file']
        file.save(os.path.join('uploads' ,request.form['filename'] + '.xlsx'))
        message = 'Yükleme başarı ile gerçekleşti...'
        try:
            #Excele ismine göre database oluşturma
            ay_secimi = request.form['filename']
            mycursor = mydb.cursor()
            sql ='''CREATE TABLE %s(
            id INT AUTO_INCREMENT PRIMARY KEY,
            police_onay varchar(255),
            police_numara varchar(255),
            police_tarih varchar(255),
            police_isim varchar(255),
            police_tutar varchar(255)
            )''' % (ay_secimi)

            mycursor.execute(sql)
            mycursor.close()
        except mysql.connector.errors.ProgrammingError:
            message = 'Bu isim zaten kullanılıyor. Lütfen farklı bir isim deneyin!'


        wb = load_workbook("{}".format(file.filename))
        ws = wb.active #Aktif çalışma sayfasını seçme | ws = wb["Şehirler"] farklı birşey için bunu uygula.

        mycursor = mydb.cursor()
        query = "INSERT INTO {} ( police_onay, police_numara, police_tarih, police_isim, police_tutar) VALUES (%s,%s,%s,%s,%s)".format(ay_secimi)

        for i in range(4 , ws.max_row+1):
            police_onay = ws.cell(i,4).value
            police_numara = ws.cell(i,2).value
            police_tarih = ws.cell(i,6).value
            police_isim = ws.cell(i,7).value
            police_tutar = ws.cell(i,8).value
            values = ( police_onay, police_numara, police_tarih, police_isim, police_tutar)
            mycursor.execute(query, values)

        mycursor.close()
        mydb.commit()
        return render_template('upload.html' , message = message)


    
    return render_template('upload.html')


if __name__ == "__main__":
    app.run(debug=True, port=8000, use_reloader=True)